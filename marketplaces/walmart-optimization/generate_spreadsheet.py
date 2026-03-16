#!/usr/bin/env python3
"""
Generate a filled Walmart MP_MAINTENANCE spreadsheet from SEO optimized data.

Reads:
  - Walmart template XLSX (downloaded from Seller Center)
  - data/seo_optimized.json (generated SEO content)
  - data/walmart_items.json (cached item data for GTINs/UPCs)

Outputs:
  - data/walmart_seo_upload.xlsx (ready to upload to Seller Center)
"""

import json
import copy
import os
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    os.system("pip3 install openpyxl")
    import openpyxl
    from openpyxl.utils import get_column_letter

# Paths
BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
TEMPLATE_PATH = Path("/Users/gabegimenes-silva/Downloads/omni-mpmaintenance-en-external-5.0.20260114-19_40_57.xlsx")
OUTPUT_PATH = DATA_DIR / "walmart_seo_upload.xlsx"

# Column mapping (1-indexed) from template inspection
COL = {
    "sku": 4,
    "specProductType": 5,
    "productIdType": 6,
    "productId": 7,
    "productName": 8,
    "brand": 9,
    "shortDescription": 10,
    "keyFeatures_0": 11,
    "keyFeatures_1": 12,
    "keyFeatures_2": 13,
    "condition": 63,
    "isProp65WarningRequired": 73,
    "light_needs": 75,
    "netContentUnit": 79,
    "netContentMeasure": 80,
    "plantCategory": 85,
    "plant_name": 88,
    "recommendedLocations": 90,
    "usda_hardiness_zone": 98,
}

# Valid light_needs values for Walmart
LIGHT_NEEDS_MAP = {
    "full sun": "Full Sun",
    "full sun to partial shade": "Full Sun to Partial Shade",
    "partial shade": "Partial Shade",
    "full shade": "Full Shade",
    "partial shade to full shade": "Partial Shade to Full Shade",
    "full sun to full shade": "Full Sun to Full Shade",
}

def normalize_light_needs(val):
    """Normalize light_needs to Walmart-accepted enum values."""
    if not val:
        return ""
    lower = val.lower().strip()
    if lower in LIGHT_NEEDS_MAP:
        return LIGHT_NEEDS_MAP[lower]
    # Try partial matching
    if "full sun" in lower and "shade" not in lower:
        return "Full Sun"
    if "partial shade" in lower and "sun" in lower:
        return "Full Sun to Partial Shade"
    if "partial shade" in lower:
        return "Partial Shade"
    if "full shade" in lower:
        return "Full Shade"
    return "Full Sun"  # default


def parse_net_content(nc_str):
    """Parse '25 lb' into (measure, unit)."""
    if not nc_str:
        return None, None
    parts = nc_str.strip().split()
    if len(parts) >= 2:
        try:
            measure = float(parts[0])
            unit = " ".join(parts[1:])
            # Normalize units
            unit_map = {
                "lb": "Pound",
                "lbs": "Pound",
                "oz": "Ounce",
                "g": "Gram",
                "kg": "Kilogram",
                "ct": "Count",
                "count": "Count",
            }
            unit = unit_map.get(unit.lower(), unit)
            return measure, unit
        except ValueError:
            pass
    return None, None


def load_walmart_items():
    """Load cached Walmart items for GTIN/UPC lookup."""
    items_path = DATA_DIR / "walmart_items.json"
    if not items_path.exists():
        print("  WARNING: walmart_items.json not found, productIdentifiers will be empty")
        return {}

    with open(items_path) as f:
        items = json.load(f)

    lookup = {}
    for item in items:
        sku = item.get("sku", "")
        if sku:
            lookup[sku] = item
    return lookup


def main():
    print("=" * 60)
    print("[GENERATE] Walmart SEO Upload Spreadsheet")
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    # Load SEO optimized data
    seo_path = DATA_DIR / "seo_optimized.json"
    with open(seo_path) as f:
        seo_items = json.load(f)
    print(f"\n  Loaded {len(seo_items)} SEO-optimized items")

    # Load Walmart items for GTIN lookup
    wm_lookup = load_walmart_items()
    print(f"  Loaded {len(wm_lookup)} Walmart items for ID lookup")

    # Load template
    print(f"\n  Loading template: {TEMPLATE_PATH.name}")
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb["Product Content And Site Exp"]

    # Data starts at row 6 (rows 1-5 are headers)
    DATA_START_ROW = 6

    # SKUs to exclude — wrong spec product type for seed template
    # Rice Hulls are "Plant Germination Trays & Kits" / "Grains", need their own template
    SKIP_SKUS = {"S-RICE-25-LB-ADDON", "S-RICE-10-LB-KIT"}

    # Fill in items
    filled = 0
    skipped = 0
    row_offset = 0

    for idx, item in enumerate(seo_items):
        sku = item["sku"]

        if sku in SKIP_SKUS:
            skipped += 1
            print(f"  SKIP: {sku} (incompatible specProductType for seed template)")
            continue

        row = DATA_START_ROW + filled
        # Get Walmart item data — use Walmart's productType, not SEO generator's
        wm_item = wm_lookup.get(sku, {})
        wm_product_type = wm_item.get("productType", "")
        # Fall back to SEO type only if Walmart has no type or "default"
        if not wm_product_type or wm_product_type == "default":
            wm_product_type = item.get("product_type", "Grass Seeds")

        # SKU (required)
        ws.cell(row=row, column=COL["sku"], value=sku)

        # Spec Product Type — must match Walmart's actual product type
        ws.cell(row=row, column=COL["specProductType"], value=wm_product_type)

        # Product Identifiers
        gtin = wm_item.get("gtin", "")
        upc = wm_item.get("upc", "")
        if gtin:
            ws.cell(row=row, column=COL["productIdType"], value="GTIN")
            ws.cell(row=row, column=COL["productId"], value=gtin)
        elif upc:
            ws.cell(row=row, column=COL["productIdType"], value="UPC")
            ws.cell(row=row, column=COL["productId"], value=upc)

        # Product Name (optimized title)
        ws.cell(row=row, column=COL["productName"], value=item.get("title", ""))

        # Brand
        ws.cell(row=row, column=COL["brand"], value="Nature's Seed")

        # Short Description (optimized description - HTML)
        desc = item.get("description", "")
        # Walmart spreadsheet accepts HTML in site description
        ws.cell(row=row, column=COL["shortDescription"], value=desc)

        # Key Features (up to 3 columns in template)
        features = item.get("key_features", [])
        for feat_idx in range(min(3, len(features))):
            col_key = f"keyFeatures_{feat_idx}"
            ws.cell(row=row, column=COL[col_key], value=features[feat_idx])

        # Attributes
        attrs = item.get("attributes", {})

        # Condition
        ws.cell(row=row, column=COL["condition"], value="New")

        # Is Prop 65 Warning Required
        ws.cell(row=row, column=COL["isProp65WarningRequired"], value="No")

        # Fill seed-specific attributes for all seed product types
        SEED_TYPES = ("Grass Seeds", "Plant Seeds", "Wildflower Seeds")
        if wm_product_type in SEED_TYPES:
            # Light Needs
            light = normalize_light_needs(attrs.get("light_needs", ""))
            if light:
                ws.cell(row=row, column=COL["light_needs"], value=light)

            # Plant Category
            plant_cat = attrs.get("plantCategory", "")
            if plant_cat:
                ws.cell(row=row, column=COL["plantCategory"], value=plant_cat)

            # Plant Name
            plant_name = attrs.get("plantName", "")
            if plant_name:
                ws.cell(row=row, column=COL["plant_name"], value=plant_name)

            # Recommended Locations
            rec_loc = attrs.get("recommendedLocations", "Outdoor")
            ws.cell(row=row, column=COL["recommendedLocations"], value=rec_loc)

            # USDA Hardiness Zone
            usda = attrs.get("usdaHardinessZone", "")
            if usda:
                ws.cell(row=row, column=COL["usda_hardiness_zone"], value=usda)

        # Net Content
        nc_str = attrs.get("netContent", "")
        if nc_str:
            measure, unit = parse_net_content(nc_str)
            if measure is not None:
                ws.cell(row=row, column=COL["netContentMeasure"], value=measure)
                ws.cell(row=row, column=COL["netContentUnit"], value=unit)

        filled += 1

    # Save
    wb.save(OUTPUT_PATH)

    print(f"\n  Filled {filled} items (skipped {skipped})")
    print(f"  Output: {OUTPUT_PATH}")
    print(f"\n  Template columns used:")
    for name, col_num in sorted(COL.items(), key=lambda x: x[1]):
        print(f"    Col {col_num:3d} ({get_column_letter(col_num):>3s}): {name}")

    print(f"\nFinished: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)


if __name__ == "__main__":
    main()
